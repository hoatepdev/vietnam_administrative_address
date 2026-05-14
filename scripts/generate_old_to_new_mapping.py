#!/usr/bin/env python3
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "admin_mapping" / "admin_mapping_old_to_new_10_25.xlsx"
OUTPUT = ROOT / "data" / "old_to_new.json"
SHEET_NAME = "admin_mapping"

REQUIRED_COLUMNS = [
    "city_id_old",
    "city_name_old",
    "district_id_old",
    "district_name_old",
    "ward_id_old",
    "ward_name_old",
    "city_id_new",
    "city_name_new",
    "ward_id_new",
    "ward_new_name",
]

ADMIN_PREFIXES = [
    "thanh pho",
    "thi tran",
    "thi xa",
    "dac khu",
    "tinh",
    "quan",
    "huyen",
    "phuong",
    "xa",
    "tp",
]


def to_string(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize(value):
    text = to_string(value).lower().strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = text.replace("đ", "d")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()

    stripped_prefix = True
    while stripped_prefix:
        stripped_prefix = False
        for prefix in ADMIN_PREFIXES:
            if text == prefix:
                return ""
            if text.startswith(prefix + " "):
                text = text[len(prefix) + 1 :].strip()
                stripped_prefix = True
                break

    return re.sub(r"\s+", " ", text).strip()


def add_index(indexes, index_name, key, row_index):
    if not key:
        return
    indexes.setdefault(index_name, {}).setdefault(key, []).append(row_index)


def main():
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    worksheet = workbook[SHEET_NAME]
    raw_headers = [to_string(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    headers = {header: idx for idx, header in enumerate(raw_headers)}

    missing_columns = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing_columns:
        raise SystemExit(f"Missing required columns: {', '.join(missing_columns)}")

    rows = []
    indexes = {
        "by_old_name_path": {},
        "by_old_province_district": {},
        "by_old_province_name": {},
        "by_old_district_name": {},
        "by_old_ward_name": {},
        "by_old_province_code": {},
        "by_old_district_code": {},
        "by_old_ward_code": {},
    }

    for excel_row in worksheet.iter_rows(min_row=2, values_only=True):
        values = {column: to_string(excel_row[headers[column]]) for column in REQUIRED_COLUMNS}
        if not any(values.values()):
            continue

        old_province_key = normalize(values["city_name_old"])
        old_district_key = normalize(values["district_name_old"])
        old_ward_key = normalize(values["ward_name_old"])
        new_province_key = normalize(values["city_name_new"])
        new_ward_key = normalize(values["ward_new_name"])

        row = {
            "old": {
                "province_code": values["city_id_old"],
                "province_name": values["city_name_old"],
                "district_code": values["district_id_old"],
                "district_name": values["district_name_old"],
                "ward_code": values["ward_id_old"],
                "ward_name": values["ward_name_old"],
            },
            "new": {
                "province_code": values["city_id_new"],
                "province_name": values["city_name_new"],
                "ward_code": values["ward_id_new"],
                "ward_name": values["ward_new_name"],
            },
            "keys": {
                "old_province_name": old_province_key,
                "old_district_name": old_district_key,
                "old_ward_name": old_ward_key,
                "old_name_path": "|".join([old_province_key, old_district_key, old_ward_key]),
                "old_province_district": "|".join([old_province_key, old_district_key]),
                "new_province_name": new_province_key,
                "new_ward_name": new_ward_key,
            },
        }

        row_index = len(rows)
        rows.append(row)

        add_index(indexes, "by_old_name_path", row["keys"]["old_name_path"], row_index)
        add_index(indexes, "by_old_province_district", row["keys"]["old_province_district"], row_index)
        add_index(indexes, "by_old_province_name", old_province_key, row_index)
        add_index(indexes, "by_old_district_name", old_district_key, row_index)
        add_index(indexes, "by_old_ward_name", old_ward_key, row_index)
        add_index(indexes, "by_old_province_code", values["city_id_old"], row_index)
        add_index(indexes, "by_old_district_code", values["district_id_old"], row_index)
        add_index(indexes, "by_old_ward_code", values["ward_id_old"], row_index)

    payload = {
        "meta": {
            "version": "10_25",
            "source": SOURCE.name,
            "sheet": SHEET_NAME,
            "row_count": len(rows),
        },
        "rows": rows,
        "indexes": indexes,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(rows)} rows to {OUTPUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
